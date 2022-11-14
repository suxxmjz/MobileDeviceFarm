import multiprocessing
from manager import Manager
from airtest.core.api import *
from ppadb.client import Client as AdbClient
import time
from multiprocessing import Event, Process
import time
import os
import pandas as pd
from datetime import datetime
from datetime import date
import openpyxl
import threading
import asyncio
import subprocess
from airtest.core.android.android import Android

class Session:
    '''
    A class to instantiate a session object.
    A session is defined as one script running on multiple devices.
    ...
    Attributes
    ------------

    scriptSession : str
        name of the script in the session
    APKsession : str
        name of the APK in the session
    packageSession : str
        name of the package in the session
    sessionID : str
        unique session ID
    devSession : list
        device to be used
    numDevs : int
        number of devices
    sessionPid : str
        process ID of session
    record : bool
        option to record device screen
    '''
    scriptSession : str = None
    APKSession : str = None
    packageSession : str = None
    sessionID : str = None
    devSession : list = []
    numDevs : int = None
    sessionPid : str = None
    sessionDuration : int = None
    record : bool = None
    uninstallAPK : bool = None

    def __init__(self, scriptSession : str, APKSession : str, numDevs : int, packageSession : str, sessionDuration : int, sessionID : str, record : str, uninstallAPK : bool):
        '''
        Constructor to set up session object.

            Parameters:
                scriptSession (str): name of script
                APKSession(str) : path of APK on computer
                numDevs(int) : number of devices
                packageSession(str) : package name
                record(str) : screen record device or no
                
        '''
        self.scriptSession = scriptSession
        self.APKSession = APKSession
        self.numDevs = numDevs
        self.packageSession = packageSession
        self.sessionID = sessionID
        self.devSession = []
        self.sessionDuration = sessionDuration
        self.uninstallAPK = uninstallAPK
        self.record = record

    def getScript(self) -> str:
        return self.scriptSession

    def getAPK(self) -> str:
        return self.APKSession
    
    def getNumDevs(self) -> int:
        return self.numDevs

    def getPackageSession(self) -> str:
        return self.packageSession
    
    def addDevice(self, theDev : str):
        self.devSession.append(theDev)
    
    def getSerials(self) -> list:
        return self.devSession
    
    def getID(self) -> str:
        return self.sessionID
    
    def setPID(self, thepid) -> None:
        self.sessionPid = thepid

    def getPID(self) -> str:
        return self.sessionPid
    
    def getDuration(self) -> int:
        return self.sessionDuration
    
    def getRecord(self) -> bool:
        return self.record
    
    def getAPKOption(self) -> bool:
        return self.uninstallAPK

    def __str__(self) -> None:
        return "Session ID:" + self.sessionID +  "  Package:" + self.packageSession + "  Number of Devices:" + str(self.numDevs)  + "  Script:" + self.scriptSession + "  APK:" + self.APKSession +  "  Devices used:" + str(self.devSession) 

    def clearSerials(self) -> None:
        self.devSession.clear() #clear all serials in devSession
        
class Runner:
    '''
    A class to create and run the sessions.

    ...
    Attributes
    -----------
    wb 
        load excel workbook
    sheet 
        load excel sheet
    writer 
        open excel writer stream
    df 
        Pandas dataframe that contains all devices
    client
        establish ADB connection to retrieve all devices
    sessionsNotRun : list
        sessions that have not been run and are still in queue
    sessionsRunning : list
        sessions that are currently running
    sessionsDone : list
        sessions that are finished
    __statuses : list
        all statuses of devices connected
    __allDevices : list
        all devices connected
    __ipAddress : list
        list of device ip addresses
    excelSave
        event instance for writing into the excel file
    '''
    wb = None
    sheet = None
    writer = None
    df = None
    client = None
    sesssionsNotRun : list = []
    sessionsRunning : list = []
    sessionsDone : list = []
    __statuses : list = []
    __allDevices : list = []
    __ipAddresses : list = []

    
  
    excelSave = Event()
    excelSave.set()
    

    def __init__(self) -> None:
        '''
        Initialize Runner class, establishes adb connection and sets devices and their statuses. 
        Also starts loop to check the queue.
        '''
        subprocess.call(("adb start-server"),shell = True)
        self.client = AdbClient()
        devices = self.client.devices(state = 'device')
        offlineDevs = self.client.devices(state = 'offline')

        for device in devices:
            self.__statuses.append('Available')
            self.__allDevices.append(device.serial)
            dev = connect_device("Android:///" + device.serial)
            set_current(device.serial)
            ip = dev.get_ip_address()
            self.__ipAddresses.append(ip)
            dev.wake()

        for device in offlineDevs:
            self.__statuses.append('Offline')
            self.__allDevices.append(device.serial)
            dev = connect_device("Android:///" + device.serial)
            set_current(device.serial)
            ip = dev.get_ip_address()
            self.__ipAddresses.append(ip)
            dev.wake()
        
            
        zipped = list(zip(self.__allDevices,self.__statuses, self.__ipAddresses))

        self.df = pd.DataFrame(zipped, columns = ['Device', 'Status','IP'])
        self.df.set_index('Device', inplace = True)
        print(self.df)
        self.df.to_excel('devices.xlsx')
        try:
            os.mkdir('DeviceHistory')
        except:
            pass
        with open('inQueue.txt', 'w') as inQ:
            inQ.write('Nothing in queue.')
            inQ.close()

        with open('activeSessions.txt', 'w') as active:
            active.write('No sessions are currently active.')
            active.close()

        with open('completedSessions.txt', 'w') as compl:
            compl.write('No sessions have been completed.')
            compl.close()

        self.loop = asyncio.new_event_loop()
        asyncio.set_event_loop(self.loop)
        threading.Thread(target = self.__startloop, args = ()).start()
    


    def __runManager(self,script : str, device : str, APK : str, package : str, ID : str, duration : int, sr : bool, removeAPK : bool) -> None:
            '''
            Calls the manager.py file and runs the program on each device.

            Parameters:
                script (str): script to run
                device(str) : device to be used
                APK(str) : APK to use
                package(str) : package to use
                ID(str) : session ID
                duration(int) : duration of how long the script will be run
                sr(bool) : screen record device
                removeAPK(bool) : whether or not to remove APK after session
            '''
            now = datetime.now() #get current time

            current_time = now.strftime("%H:%M:%S")
            f = open('DeviceHistory/' + device + '.txt', 'a')
            
            m1 = Manager(script,device, APK, package, ID, sr)
            m1.checkInputs()
            m1.installAPK()
            time.sleep(20)
            m1.runScript()
            
            time.sleep(duration)
            m1.kill()
            m1.stopAPK()
            if removeAPK == True:
                m1.uninstallAPK()
            else:
                pass
            todays_date = date.today()
            f.write(str(todays_date) + ", " + current_time + " : Device " + device + " ran " +script ) #write into device history file
            f.write('\n')
            f.close()
            

    def __updateStatus(self,dev : str) -> None:
        '''
        Updates status of device to available at the end of program.

        Parameters: 
            dev(str) : device to be updated
        '''
        n =2 #counter to keep track of excel sheet rows
        rowCount = self.sheet.max_row #total number of rows in excel
        while n <= rowCount:
            cell = self.sheet.cell(row = n, column = 1) #status cell of excel sheet to check
            if cell.value == dev:
                finishedDev = self.sheet.cell(row = n, column = 1) #the device cell which will be changed
                self.df.at[finishedDev.value, 'Status'] = 'Available'
                self.__saveToExcel()
                break
                        
            else:
                n +=1


    def __saveToExcel(self) -> None:
        '''
        Saves to excel.
        '''
        self.excelSave.wait()
        self.excelSave.clear()
        self.df.to_excel(self.writer, sheet_name = 'Sheet1', encoding= 'utf-8')
        self.writer.save()
        time.sleep(1)
        self.excelSave.set()
        
    def __createSession(self,newSession : Session) -> None:
        '''
        Creates a new session.

        Parameters:
            newSession(object) : the session object
        '''

        self.wb = openpyxl.load_workbook('devices.xlsx')
        self.sheet = self.wb.active
        self.writer = pd.ExcelWriter('devices.xlsx')


        numberOfDev = newSession.getNumDevs()
        n = 2 #keep track of excel sheet rows
        i = 0 #keep track of number of devices added
        devicesToRemove = [] # temp array that stores devices to be used, if enough aren't met, then remove these devices
        deviceComplete : bool = None #checks if enough devices were found

        while i < numberOfDev:
                if self.df.empty:
                    deviceComplete = False
                    break

                cell = self.sheet.cell(row = n, column = 2)
                if cell.value == 'Available':
                    deviceNumber = self.sheet.cell(row = n, column = 1)
                    newSession.addDevice(deviceNumber.value)
                    self.df.at[deviceNumber.value, 'Status'] = 'Taken'
                    
                    n+=1
                    i += 1
                    if i == numberOfDev:
                        deviceComplete = True
        
                else:
                    n +=1
                    j = n-1
                    if j == self.sheet.max_row:
                        deviceComplete = False
                        devicesToRemove = newSession.getSerials()
                        for d in devicesToRemove:
                            self.df.at[d, 'Status'] = 'Available'
                            
                        newSession.clearSerials()
                        break
        self.__saveToExcel()
        return deviceComplete
            
        
    def __runSession(self, newSession : Session) -> None:
        '''
        Runs the sessions, calls on runManager. Parameters are fetched from the session object.
        Parameters:
            newSession(object) : session object to use
        '''

        deviceSerials = newSession.getSerials() #devices to be used
        nameOfScript =  newSession.getScript() #name of script
        APKPath = newSession.getAPK() #path of APK on computer
        nameOfPackage = newSession.getPackageSession() #package name
        sessID = newSession.getID() #session ID
        sessDur = newSession.getDuration() #how long the session script will run for
        sr = newSession.getRecord() #screenrecord or no
        remAPK = newSession.getAPKOption() #remove apk after session or not
        

        for dev in deviceSerials:
            p = Process(target = self.__runManager, args = (nameOfScript,dev,APKPath,nameOfPackage, sessID, sessDur, sr, remAPK))
            p.start()
        p.join()

    def takeInput(self, nameOfScript : str, APKPath : str,nameOfPackage : str,  numberOfDev : int, duration : int , id : str, recordScreen : bool, removeAPK : bool) -> None:
        '''
        Takes in input to run.
        Parameters:
            nameOfScript(str) : script name
            APKPath(str) : path to apk
            nameOfPackage(str) : package name
            numberOfDev(int) : number of devices to be used
            duration(int) : duration of how long the script will be run
            id(str) : session ID
            recordScreen(bool) : yes or no for screenrecord
            removeAPK(bool) : yes or no to uninstall APK after session
        '''

        
        newSession = Session(nameOfScript, APKPath, numberOfDev,nameOfPackage, duration, id, recordScreen, removeAPK)

        self.sesssionsNotRun.append(newSession)

    def __updateQueue(self,s : Session, p : Process = None) -> None:
        '''
        Updates queue.
        Parameters:
            s(Session) : the session to add and remove from the queues
            p(Process) : the process to wait for
        '''
        if p == None:
            pass
        else:
            p.join()
        theDevs = s.getSerials() #gets serials from object
        for each in theDevs:
            self.__updateStatus(each)
        self.sessionsRunning.remove(s)
        self.sessionsDone.append(s)


    async def __runQueue(self) -> None:
            '''
            Checks the queue, runs if session is created with enough devices.
            '''
            if len(self.sesssionsNotRun) > 0 :
                theSession : Session = None
                theSession = self.sesssionsNotRun[0] #gets first session from queue
                self.sesssionsNotRun.remove(theSession)
                #time.sleep(5)
                deviceComplete = self.__createSession(theSession)

                if deviceComplete == True:
                    self.sessionsRunning.append(theSession)
                    p = multiprocessing.Process(target = self.__runSession, args= (theSession,))
                    process = multiprocessing.current_process()
                    theSession.setPID(process.pid)
                    p.start()
                    threading.Thread(target = self.__updateQueue, args = (theSession,p)).start()
                    
                elif deviceComplete == False:
                        self.sesssionsNotRun.append(theSession)
                        

            self.__queueStatus()


    async def __callQueue(self) -> None:
        '''
        Calls queue every 5 seconds.
        '''
        while True:
            time.sleep(5)
            await self.__runQueue()
                

    def __startloop(self) -> None:
        '''
        Starts event loop.
        '''
        self.loop.run_until_complete(self.__callQueue())
        self.loop.run_forever

    def __queueStatus(self) -> None:
        ''''
        Prints queues and shows which sessions are not running, running and completed.
        '''
        with open('inQueue.txt', 'w') as f:
            if not self.sesssionsNotRun:
                f.write("Nothing in queue.")

            for each in self.sesssionsNotRun:
                f.write(str(each))
                f.write("\n")

        with open('activeSessions.txt', 'w') as a:
            if not self.sessionsRunning:
                a.write("No sessions are currently active.")

            for each in self.sessionsRunning:
                a.write(str(each))
                a.write("\n")
        with open('completedSessions.txt', 'w') as c:
            if not self.sessionsDone:
                c.write("No sessions have been completed.")

            for each in self.sessionsDone:
                c.write(str(each))
                c.write("\n")



#beginProgram = Runner()



#beginProgram.takeInput('Android/instagram.air', 'Instagram.apk','com.instagram.android', 2, 80, '234435' )
#time.sleep(5)
#beginProgram.takeInput('Android/reddit.air', 'reddit.apk','com.reddit.frontpage', 2, 60, '23d4323' )
#time.sleep(10)
#beginProgram.takeInput('Android/facebook.air', 'Facebook.apk','com.facebook.katana', 2, 110, '2777' )
